import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .aggregation import act_vs_plan_for_date_range, act_vs_plan_for_shift_instance, daily_production_report_rows
from .availability import availability_utilization

# Shared styling — applied to every exported sheet so "download the report"
# consistently produces something a Supervisor can hand off or print
# directly, not a plain data dump: bold white-on-blue headers, thousands-
# separated/1-decimal numbers, a real "%" suffix on variance/availability
# figures (these values already arrive pre-multiplied by 100, e.g. 93.68
# meaning "93.68%" — a true Excel percent format would divide by 100 again
# and show "0.0%"), thin borders, frozen header row(s), and auto-sized
# columns instead of a fixed guess.
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_THIN = Side(style="thin", color="B7B7B7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center")
_NUM_FMT = "#,##0.0"
_PCT_FMT = '0.0"%"'


def _style_header_row(ws, row, num_cols, start_col=1):
    for col in range(start_col, start_col + num_cols):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = _BORDER


def _autosize_columns(ws, min_width=9, max_width=42):
    """openpyxl has no built-in auto-fit — approximate it from the widest
    rendered value in each column, since every sheet here was previously
    left at Excel's default ~8.4-character column width regardless of
    content (headers/values got clipped or the column was needlessly wide).
    """
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, length in widths.items():
        ws.column_dimensions[col].width = min(max(length + 2, min_width), max_width)


def _write_act_vs_plan_rows(ws, rows):
    headers = ["Section", "Parameter", "UOM", "Act", "Plan", "Var", "%Var"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for row in rows:
        ws.append(
            [
                row["section_name"],
                row["parameter_name"],
                row["uom"] or "",
                float(row["act"]),
                float(row["plan"]) if row["plan"] is not None else None,
                float(row["var"]) if row["var"] is not None else None,
                row["pct_var"],
            ]
        )

    last_row = ws.max_row
    _border_range(ws, 1, last_row, 1, len(headers))
    for r in range(2, last_row + 1):
        for col_letter in ("D", "E", "F"):
            ws[f"{col_letter}{r}"].number_format = _NUM_FMT
        ws[f"G{r}"].number_format = _PCT_FMT

    ws.freeze_panes = "A2"
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    _autosize_columns(ws)


def build_shift_report_xlsx(shift_instance):
    """Formats close to the source Daily Production Report layout: one
    Act/Plan/Var/%Var row per section+parameter for the shift.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Shift {shift_instance.date}"
    _write_act_vs_plan_rows(ws, act_vs_plan_for_shift_instance(shift_instance))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_daily_report_xlsx(site, date):
    """Same Act/Plan/Var/%Var shape as the shift report, widened to every
    shift instance on one calendar day — matched against PlanTarget's
    PERIOD_DAY figures for that date.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily {date}"
    rows = act_vs_plan_for_date_range(site.id, date, date, plan_period_type="day", plan_period_date=date)
    _write_act_vs_plan_rows(ws, rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# Header band colors matching the source spreadsheet's blue/orange/green
# convention (Day Shift = orange, Night Shift = green, Daily Total/MTD =
# blue, Availabilities = light green) — a fixed, non-brand-neutral palette
# on purpose, since this is reproducing a specific existing report a mine
# already circulates, not a new dashboard visualization.
_FILL_BLUE = PatternFill("solid", fgColor="9DC3E6")
_FILL_ORANGE = PatternFill("solid", fgColor="F4B183")
_FILL_GREEN = PatternFill("solid", fgColor="A9D18E")
_FILL_LIGHT_GREEN = PatternFill("solid", fgColor="C6E0B4")
_BOLD = Font(bold=True)


def _write_group_header(ws, row, start_col, span, label, fill):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + span - 1)
    cell = ws.cell(row=row, column=start_col, value=label)
    cell.font = _BOLD
    cell.alignment = _CENTER
    for col in range(start_col, start_col + span):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = _BORDER


def build_daily_production_report_xlsx(site, date):
    """Reproduces the source "Daily Production Report" spreadsheet's exact
    layout: one row per parameter, with Act/Plan/%Var column-groups for
    each shift on `date`, plus Daily Total and Month-to-Date column-groups
    (Act/Plan/Var/%Var), plus an Availabilities/Utilization block by
    machine type. Data comes from daily_production_report_rows (per-shift
    + daily + MTD, collapsed across sections to match this report's grain)
    and the existing availability_utilization service — both already used
    elsewhere, just combined into one sheet here for the first time.
    """
    import datetime as dt

    date = dt.date.fromisoformat(date) if isinstance(date, str) else date
    report = daily_production_report_rows(site.id, date)
    shift_names = report["shift_names"]
    rows = report["rows"]
    availability = availability_utilization(site.id, date)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily {date}"

    ws.cell(row=1, column=1, value=site.name).font = _TITLE_FONT
    ws.cell(row=1, column=6, value="Day")
    ws.cell(row=1, column=7, value=date.day).font = Font(bold=True, color="FF0000")
    ws.cell(row=2, column=1, value="Daily Production Report").font = _BOLD

    header_row = 4
    subheader_row = 5
    data_start_row = 6

    ws.cell(row=header_row, column=1, value="Parameter").font = _BOLD
    ws.cell(row=header_row, column=2, value="UOM").font = _BOLD
    ws.cell(row=subheader_row, column=1).border = _BORDER
    ws.cell(row=subheader_row, column=2).border = _BORDER

    col = 3
    shift_col_start = {}
    for shift_name in shift_names:
        shift_col_start[shift_name] = col
        fill = _FILL_ORANGE if col == 3 else _FILL_GREEN
        _write_group_header(ws, header_row, col, 3, f"{shift_name} Shift", fill)
        for i, label in enumerate(["Act", "Plan", "%Var"]):
            cell = ws.cell(row=subheader_row, column=col + i, value=label)
            cell.font = _BOLD
            cell.alignment = _CENTER
            cell.border = _BORDER
        col += 3

    daily_total_col = col
    _write_group_header(ws, header_row, col, 4, "Daily Total", _FILL_BLUE)
    for i, label in enumerate(["Act", "Plan", "Var", "%Var"]):
        cell = ws.cell(row=subheader_row, column=col + i, value=label)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER
    col += 4

    mtd_col = col
    _write_group_header(ws, header_row, col, 4, "Month to Date", _FILL_BLUE)
    for i, label in enumerate(["Act", "Plan", "Var", "%Var"]):
        cell = ws.cell(row=subheader_row, column=col + i, value=label)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER
    col += 4

    avail_col = col + 1  # one spacer column between the Act/Plan blocks and Availabilities
    _write_group_header(ws, header_row, avail_col, 4, "Availabilities", _FILL_LIGHT_GREEN)
    for i, label in enumerate(["Machine", "Day", "Night", "Average"]):
        cell = ws.cell(row=subheader_row, column=avail_col + i, value=label)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER

    pct_var_cols = set()
    for shift_name, start in shift_col_start.items():
        pct_var_cols.add(start + 2)
    pct_var_cols.add(daily_total_col + 3)
    pct_var_cols.add(mtd_col + 3)
    avail_pct_cols = {avail_col + 1, avail_col + 2, avail_col + 3}

    for r, row in enumerate(rows):
        excel_row = data_start_row + r
        ws.cell(row=excel_row, column=1, value=row["parameter_name"])
        ws.cell(row=excel_row, column=2, value=row["uom"] or "")
        for shift_name, start in shift_col_start.items():
            block = row["by_shift"].get(shift_name, {})
            ws.cell(row=excel_row, column=start, value=float(block.get("act", 0)))
            ws.cell(row=excel_row, column=start + 1, value=float(block["plan"]) if block.get("plan") is not None else None)
            ws.cell(row=excel_row, column=start + 2, value=block.get("pct_var"))
        dt_block = row["daily_total"]
        ws.cell(row=excel_row, column=daily_total_col, value=float(dt_block["act"]))
        ws.cell(row=excel_row, column=daily_total_col + 1, value=float(dt_block["plan"]) if dt_block["plan"] is not None else None)
        ws.cell(row=excel_row, column=daily_total_col + 2, value=float(dt_block["var"]) if dt_block["var"] is not None else None)
        ws.cell(row=excel_row, column=daily_total_col + 3, value=dt_block["pct_var"])
        mtd_block = row["mtd"]
        ws.cell(row=excel_row, column=mtd_col, value=float(mtd_block["act"]))
        ws.cell(row=excel_row, column=mtd_col + 1, value=float(mtd_block["plan"]) if mtd_block["plan"] is not None else None)
        ws.cell(row=excel_row, column=mtd_col + 2, value=float(mtd_block["var"]) if mtd_block["var"] is not None else None)
        ws.cell(row=excel_row, column=mtd_col + 3, value=mtd_block["pct_var"])

    for r, machine_row in enumerate(availability):
        excel_row = data_start_row + r
        by_shift_name = {b["shift_name"]: b for b in machine_row["by_shift"]}
        day_pct = next((b["availability_pct"] for name, b in by_shift_name.items() if name in shift_names[:1]), None)
        night_pct = next((b["availability_pct"] for name, b in by_shift_name.items() if name in shift_names[1:2]), None)
        ws.cell(row=excel_row, column=avail_col, value=machine_row["machine_type_name"])
        ws.cell(row=excel_row, column=avail_col + 1, value=day_pct)
        ws.cell(row=excel_row, column=avail_col + 2, value=night_pct)
        ws.cell(row=excel_row, column=avail_col + 3, value=machine_row["average"]["availability_pct"])

    last_row = data_start_row + max(len(rows), len(availability)) - 1
    last_col = avail_col + 3
    if last_row >= data_start_row:
        _border_range(ws, data_start_row, last_row, 1, last_col)
        for r in range(data_start_row, last_row + 1):
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell.value, (int, float)):
                    continue
                cell.number_format = _PCT_FMT if c in pct_var_cols or c in avail_pct_cols else _NUM_FMT

    ws.freeze_panes = ws.cell(row=data_start_row, column=3).coordinate
    _autosize_columns(ws, min_width=10)
    ws.column_dimensions["A"].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_mtd_report_xlsx(site, year, month):
    """Month-to-date Act/Plan/Var/%Var: every shift instance from the 1st of
    the month through today (or the month's end, if generated later),
    matched against PlanTarget's PERIOD_MONTH figure keyed to the month's
    first day (the seed data / admin UI convention for a monthly target).
    """
    import datetime as dt

    month_start = dt.date(year, month, 1)
    today = dt.date.today()
    month_end = (
        today if (today.year, today.month) == (year, month) else dt.date(year, month, 28) + dt.timedelta(days=4)
    )
    if month_end.month != month:
        # rolled into next month via the day-28-plus-4 trick — clamp back
        month_end = month_end.replace(day=1) - dt.timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = f"MTD {year}-{month:02d}"
    rows = act_vs_plan_for_date_range(
        site.id, month_start, month_end, plan_period_type="month", plan_period_date=month_start
    )
    _write_act_vs_plan_rows(ws, rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
