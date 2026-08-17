from datetime import timedelta

import pytest
from openpyxl import load_workbook

from dashboard.services.aggregation import act_vs_plan_for_shift_instance
from dashboard.services.export import (
    build_daily_production_report_xlsx,
    build_daily_report_xlsx,
    build_mtd_report_xlsx,
    build_shift_report_xlsx,
)
from dashboard.services.hourly_machine_status import hourly_machine_status
from entries.models import BreakdownLog, ParameterValue, ProductionEntry
from machines.models import Machine
from masterdata.models import UOM, Parameter
from shiftmgmt.services import get_or_create_open_instance, time_slots_for_instance


@pytest.mark.django_db
def test_average_aggregation_parameter_is_averaged_not_summed(
    machines, sections, all_day_shifts, django_user_model
):
    """Regression test: a Parameter.aggregation="average" parameter (e.g.
    Machine Availability %) must roll up hourly readings by averaging them,
    not summing — three hourly readings of 99/99/100% must produce ~99.33,
    not the meaningless "298" a plain Sum would give.
    """
    m_a, _ = machines
    sec_a, _ = sections
    operator = django_user_model.objects.create_user(username="op1", password="pass12345")
    uom = UOM.objects.create(name="Percent", abbreviation="%")
    parameter = Parameter.objects.create(
        name="Machine Availability",
        code="machine-availability",
        uom=uom,
        scope=Parameter.SCOPE_MACHINE,
        data_type=Parameter.DATA_TYPE_NUMBER,
        aggregation=Parameter.AGGREGATION_AVERAGE,
    )
    parameter.applicable_machine_types.add(m_a.machine_type)

    instance = get_or_create_open_instance(m_a.site)
    for slot_index, value in enumerate([99, 99, 100]):
        entry = ProductionEntry.objects.create(
            shift_instance=instance,
            site=m_a.site,
            section=sec_a,
            machine=m_a,
            entry_type=ProductionEntry.ENTRY_TYPE_HOURLY,
            slot_index=slot_index,
            operator=operator,
            recorded_by=operator,
        )
        ParameterValue.objects.create(production_entry=entry, parameter=parameter, value_number=value)

    results = act_vs_plan_for_shift_instance(instance)
    row = next(r for r in results if r["parameter"] == parameter.id)
    assert float(row["act"]) == pytest.approx(99.333, abs=0.01)


@pytest.mark.django_db
def test_sum_aggregation_parameter_still_sums(machines, sections, all_day_shifts, django_user_model):
    """The default/unset aggregation ("sum") must keep behaving exactly as
    before this field existed — additive quantities like tonnes hauled are
    genuinely a total across the shift's hourly entries."""
    m_a, _ = machines
    sec_a, _ = sections
    operator = django_user_model.objects.create_user(username="op1", password="pass12345")
    parameter = Parameter.objects.create(
        name="Tonnes Hauled", code="tonnes-hauled-sum-test", scope=Parameter.SCOPE_MACHINE,
        data_type=Parameter.DATA_TYPE_NUMBER,
    )
    parameter.applicable_machine_types.add(m_a.machine_type)

    instance = get_or_create_open_instance(m_a.site)
    for slot_index, value in enumerate([40, 45, 50]):
        entry = ProductionEntry.objects.create(
            shift_instance=instance,
            site=m_a.site,
            section=sec_a,
            machine=m_a,
            entry_type=ProductionEntry.ENTRY_TYPE_HOURLY,
            slot_index=slot_index,
            operator=operator,
            recorded_by=operator,
        )
        ParameterValue.objects.create(production_entry=entry, parameter=parameter, value_number=value)

    results = act_vs_plan_for_shift_instance(instance)
    row = next(r for r in results if r["parameter"] == parameter.id)
    assert row["act"] == 135


@pytest.mark.django_db
def test_hourly_machine_status_marks_only_the_slots_a_breakdown_overlaps(
    machines, sections, all_day_shifts, django_user_model
):
    """Regression test for the Availability & Breakdown Report grid: a
    BreakdownLog spanning the middle of one hour slot must mark that slot
    (and only that slot) as down, with the reason surfaced and the
    running-count for that slot dropping to reflect it."""
    m_a, _ = machines
    sec_a, _ = sections
    operator = django_user_model.objects.create_user(username="op3", password="pass12345")
    instance = get_or_create_open_instance(m_a.site)

    slots = time_slots_for_instance(instance)
    _, slot1_start, slot1_end = slots[1]
    BreakdownLog.objects.create(
        shift_instance=instance,
        site=m_a.site,
        section=sec_a,
        machine=m_a,
        start_at=slot1_start + timedelta(minutes=5),
        end_at=slot1_end - timedelta(minutes=5),
        description="Hydraulic hose burst",
        operator=operator,
        recorded_by=operator,
    )

    results = hourly_machine_status(instance)
    group = next(g for g in results if g["machine_type"] == m_a.machine_type_id)
    row = next(m for m in group["machines"] if m["machine"] == m_a.id)

    down_slots = {c["slot_index"] for c in row["cells"] if not c["ok"]}
    assert down_slots == {1}
    assert row["cells"][1]["reason"] == "Hydraulic hose burst"
    assert group["running_by_slot"][0] == 1
    assert group["running_by_slot"][1] == 0


@pytest.mark.django_db
def test_hourly_machine_status_excludes_crushers(machines, sections, all_day_shifts, django_user_model):
    """Crusher breakdowns are tracked through the Crushing & Breakdowns
    module's own hourly checklist/matrix, not BreakdownLog — a Crusher-type
    Machine must never show up on this general-fleet grid."""
    m_a, _ = machines
    from masterdata.models import MachineType

    crusher_type = MachineType.objects.create(name="Crusher", code="cru")
    Machine.objects.create(site=m_a.site, machine_type=crusher_type, fleet_number="CRU-1")

    instance = get_or_create_open_instance(m_a.site)
    results = hourly_machine_status(instance)

    assert all(g["machine_type_name"] != "Crusher" for g in results)


@pytest.fixture
def act_vs_plan_data(machines, sections, all_day_shifts, django_user_model):
    """One hourly entry + a same-shift PlanTarget, so every export builder
    below has at least one real Act/Plan/Var/%Var row to format."""
    from planning.models import PlanTarget

    m_a, _ = machines
    sec_a, _ = sections
    operator = django_user_model.objects.create_user(username="op_export", password="pass12345")
    parameter = Parameter.objects.create(
        name="Tonnes Hauled", code="tonnes-hauled-export-test", scope=Parameter.SCOPE_MACHINE,
        data_type=Parameter.DATA_TYPE_NUMBER,
    )
    parameter.applicable_machine_types.add(m_a.machine_type)

    from entries.services import resolve_slot_datetimes

    instance = get_or_create_open_instance(m_a.site)
    slot_start_at, slot_end_at = resolve_slot_datetimes(instance, 0)
    entry = ProductionEntry.objects.create(
        shift_instance=instance, site=m_a.site, section=sec_a, machine=m_a,
        entry_type=ProductionEntry.ENTRY_TYPE_HOURLY, slot_index=0,
        slot_start_at=slot_start_at, slot_end_at=slot_end_at,
        operator=operator, recorded_by=operator,
    )
    ParameterValue.objects.create(production_entry=entry, parameter=parameter, value_number=55)
    PlanTarget.objects.create(
        parameter=parameter, site=m_a.site, section=sec_a,
        period_type=PlanTarget.PERIOD_SHIFT, shift_instance=instance, target_value=100,
    )
    return instance, m_a.site


@pytest.mark.django_db
def test_shift_report_xlsx_is_formatted(act_vs_plan_data):
    instance, _site = act_vs_plan_data
    wb = load_workbook(build_shift_report_xlsx(instance))
    ws = wb.active

    header = ws["A1"]
    assert header.font.bold is True
    assert header.fill.fgColor.rgb == "001F4E78"
    assert ws["D2"].number_format == "#,##0.0"  # Act
    assert ws["G2"].number_format == '0.0"%"'  # %Var
    assert ws.freeze_panes == "A2"
    assert ws.column_dimensions["B"].width > 8  # "Parameter" header auto-sized, not left at default


@pytest.mark.django_db
def test_daily_report_xlsx_is_formatted(act_vs_plan_data):
    instance, site = act_vs_plan_data
    wb = load_workbook(build_daily_report_xlsx(site, instance.date))
    ws = wb.active
    assert ws["A1"].font.bold is True
    assert ws["D2"].number_format == "#,##0.0"


@pytest.mark.django_db
def test_mtd_report_xlsx_is_formatted(act_vs_plan_data):
    instance, site = act_vs_plan_data
    wb = load_workbook(build_mtd_report_xlsx(site, instance.date.year, instance.date.month))
    ws = wb.active
    assert ws["A1"].font.bold is True
    assert ws["D2"].number_format == "#,##0.0"


@pytest.mark.django_db
def test_daily_production_report_xlsx_is_formatted(act_vs_plan_data):
    instance, site = act_vs_plan_data
    wb = load_workbook(build_daily_production_report_xlsx(site, instance.date))
    ws = wb.active

    # Row 4/5 are the merged group headers + column sub-headers; data starts row 6.
    assert ws["A4"].font.bold is True
    assert ws["C4"].fill.fgColor.rgb in {"00F4B183", "00A9D18E"}  # Day/Night shift band color
    assert ws["C6"].number_format == "#,##0.0"  # first shift's Act column
    assert ws.freeze_panes == "C6"


@pytest.mark.django_db
def test_hourly_curve_includes_flat_per_hour_act_and_target(act_vs_plan_data):
    """Regression test: the Hourly Curve chart needs the raw per-hour Act
    and a flat per-hour target (shift target evenly split across slots —
    e.g. a 500-tonne/8-hour shift target is 62.5 tonnes/hour, not a ramp),
    not just the running cumulative totals — the bars half of the source
    report's "Hourly Tonnes / Hourly Target" bars-plus-lines combo chart,
    which the API previously had no field for at all."""
    from dashboard.services.aggregation import hourly_curve
    from shiftmgmt.services import time_slots_for_instance

    instance, site = act_vs_plan_data
    sec_a = instance.production_entries.first().section
    parameter = instance.production_entries.first().values.first().parameter

    slots = time_slots_for_instance(instance)
    curve = hourly_curve(instance, sec_a.id, parameter.id)

    assert len(curve) == len(slots)
    slot0 = next(p for p in curve if p["slot_index"] == 0)
    assert slot0["act"] == 55
    assert float(slot0["target"]) == pytest.approx(100 / len(slots))
    assert slot0["cumulative_act"] == 55
    assert float(slot0["cumulative_target"]) == pytest.approx(100 / len(slots))

    slot1 = next(p for p in curve if p["slot_index"] == 1)
    assert slot1["act"] == 0
    assert float(slot1["target"]) == pytest.approx(100 / len(slots))
    assert float(slot1["cumulative_target"]) == pytest.approx(2 * 100 / len(slots))


@pytest.mark.django_db
def test_general_fleet_mttr_averages_fixed_breakdowns_and_reports_the_plan_target(
    machines, sections, all_day_shifts, django_user_model
):
    """Only breakdowns the Artisan workflow has actually marked fixed (or
    confirmed) count toward MTTR — an unresolved one has no repair time
    yet. The relevant Plan Target for the "mttr-minutes" parameter (seeded
    by masterdata migration 0008) comes back alongside the actual."""
    from datetime import date

    from django.utils import timezone

    from dashboard.services.mttr import general_fleet_mttr
    from entries.models import BreakdownLog
    from masterdata.models import Parameter
    from planning.models import PlanTarget
    from shiftmgmt.services import get_or_create_open_instance

    m_a, _ = machines
    sec_a, _ = sections
    operator = django_user_model.objects.create_user(username="mttr_op", password="pass12345")
    instance = get_or_create_open_instance(m_a.site)
    today = date.today()

    for minutes, repair_status in ((30, BreakdownLog.REPAIR_FIXED), (90, BreakdownLog.REPAIR_CONFIRMED)):
        start = timezone.now()
        BreakdownLog.objects.create(
            shift_instance=instance, site=m_a.site, section=sec_a, machine=m_a,
            start_at=start, end_at=start + timedelta(minutes=minutes),
            repair_status=repair_status, operator=operator, recorded_by=operator,
        )
    # Still open — must not be counted.
    BreakdownLog.objects.create(
        shift_instance=instance, site=m_a.site, section=sec_a, machine=m_a,
        start_at=timezone.now(), repair_status=BreakdownLog.REPAIR_REPORTED,
        operator=operator, recorded_by=operator,
    )

    parameter = Parameter.objects.get(code="mttr-minutes")
    PlanTarget.objects.create(
        parameter=parameter, site=m_a.site, section=sec_a,
        period_type=PlanTarget.PERIOD_DAY, period_date=today, target_value=45,
    )

    result = general_fleet_mttr(m_a.site_id, sec_a.id, today - timedelta(days=1), today + timedelta(days=1))
    assert result["breakdown_count"] == 2
    assert float(result["actual_mttr_minutes"]) == pytest.approx(60)
    assert float(result["target_mttr_minutes"]) == pytest.approx(45)


@pytest.mark.django_db
def test_production_summary_groups_totals_by_operator_and_by_machine(
    machines, sections, all_day_shifts, django_user_model
):
    """"How many tonnes did this operator produce" — group_by="operator"
    sums across whichever machines they used; group_by="machine" sums
    across whichever operators ran it. Same underlying entries, two
    different cuts of the same total."""
    from datetime import date

    from dashboard.services.production_summary import production_summary
    from entries.services import resolve_slot_datetimes

    m_a, _ = machines
    sec_a, _ = sections
    op1 = django_user_model.objects.create_user(username="summary_op1", password="pass12345")
    op2 = django_user_model.objects.create_user(username="summary_op2", password="pass12345")
    parameter = Parameter.objects.create(
        name="Tonnes Hauled", code="tonnes-hauled-summary-test", scope=Parameter.SCOPE_MACHINE,
        data_type=Parameter.DATA_TYPE_NUMBER,
    )
    parameter.applicable_machine_types.add(m_a.machine_type)

    instance = get_or_create_open_instance(m_a.site)
    for slot_index, (operator, value) in enumerate([(op1, 40), (op1, 20), (op2, 15)]):
        slot_start_at, slot_end_at = resolve_slot_datetimes(instance, slot_index)
        entry = ProductionEntry.objects.create(
            shift_instance=instance, site=m_a.site, section=sec_a, machine=m_a,
            entry_type=ProductionEntry.ENTRY_TYPE_HOURLY, slot_index=slot_index,
            slot_start_at=slot_start_at, slot_end_at=slot_end_at,
            operator=operator, recorded_by=operator,
        )
        ParameterValue.objects.create(production_entry=entry, parameter=parameter, value_number=value)

    today = date.today()
    by_operator = production_summary(m_a.site_id, parameter.id, today, today, "operator")
    totals = {row["group"]: float(row["act"]) for row in by_operator}
    assert totals[op1.id] == pytest.approx(60)
    assert totals[op2.id] == pytest.approx(15)

    by_machine = production_summary(m_a.site_id, parameter.id, today, today, "machine")
    assert len(by_machine) == 1
    assert by_machine[0]["group"] == m_a.id
    assert float(by_machine[0]["act"]) == pytest.approx(75)
